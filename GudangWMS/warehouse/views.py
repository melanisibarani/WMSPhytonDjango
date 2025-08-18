# warehouse/views.py
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Case, When, IntegerField, F, Value
from .models import Barang, Transaksi

# -----------------------------
# 1) Barang Masuk
# -----------------------------
from django.core.paginator import Paginator

def barang_masuk(request):
    if request.method == "POST":
        kode_barang = request.POST.get("kode_barang")
        qty = request.POST.get("qty")

        if kode_barang and qty:
            barang, created = Barang.objects.get_or_create(
                kode_barang=kode_barang,
                defaults={"nama_barang": f"Barang {kode_barang}"}
            )
            Transaksi.objects.create(
                item=barang,
                jenis=Transaksi.IN,
                qty=int(qty),
                created_by=request.user
            )

    qs = Transaksi.objects.filter(jenis=Transaksi.IN).order_by("-created_at")

    # Pagination
    page_number = request.GET.get("page", 1)
    paginator = Paginator(qs, 10)  # tampilkan 10 per halaman
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "warehouse/barang_masuk.html",
        {
            "masuk_list": page_obj  # lempar page_obj ke template
        }
    )


# -----------------------------
# 2) Barang Keluar
# -----------------------------
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404

def barang_keluar(request):
    if request.method == "POST":
        kode_barang = request.POST.get("kode_barang")
        qty = request.POST.get("qty")

        if kode_barang and qty:
            barang = get_object_or_404(Barang, kode_barang=kode_barang)
            Transaksi.objects.create(
                item=barang,
                jenis=Transaksi.OUT,
                qty=int(qty),
                created_by=request.user
            )

    qs = Transaksi.objects.filter(jenis=Transaksi.OUT).order_by("-created_at")

    # Pagination
    page_number = request.GET.get("page", 1)
    paginator = Paginator(qs, 10)  # tampil 10 per halaman
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "warehouse/barang_keluar.html",
        {
            "keluar_list": page_obj  # lempar page_obj
        }
    )



# -----------------------------
# 3) Stok List (Total Qty)
# -----------------------------
def stok_list(request):
    stock = (
        Barang.objects.annotate(
            total_masuk=Sum(
                Case(When(transactions__jenis=Transaksi.IN, then="transactions__qty"),
                     default=Value(0), output_field=IntegerField())
            ),
            total_keluar=Sum(
                Case(When(transactions__jenis=Transaksi.OUT, then="transactions__qty"),
                     default=Value(0), output_field=IntegerField())
            ),
        )
        .annotate(total_qty=F("total_masuk") - F("total_keluar"))
    )

    return render(request, "warehouse/stok_list.html", {"stock": stock})


# -----------------------------
# 4) Kartu Stok (History per barang)
# -----------------------------
# views.py
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from .models import Barang, Transaksi

def kartu_stok(request, item_id):
    barang = get_object_or_404(Barang, id=item_id)
    transactions = Transaksi.objects.filter(item=barang).order_by("created_at")

    data = []
    total = 0
    for trx in transactions:
        if trx.jenis == "IN":
            qty_in = trx.qty
            qty_out = 0
            total += trx.qty
        else:
            qty_in = 0
            qty_out = trx.qty
            total -= trx.qty

        data.append({
            "tanggal": trx.created_at,
            "item": trx.item,
            "qty_in": qty_in,
            "qty_out": qty_out,
            "total": total,
        })

    # 🔹 Tambahin pagination
    paginator = Paginator(data, 10)  # 10 baris per halaman
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "warehouse/kartu_stok.html", {
        "barang": barang,
        "kartu_stok": page_obj,  # pass page_obj ke template
    })

from django.shortcuts import render, get_object_or_404

masuk_list = Transaksi.objects.filter(jenis=Transaksi.IN).order_by("-created_at")
keluar_list = Transaksi.objects.filter(jenis=Transaksi.OUT).order_by("-created_at")


from django.shortcuts import render, get_object_or_404
import qrcode, base64
from io import BytesIO
from .models import Transaksi

def print_keluar(request, pk):
    transaksi = get_object_or_404(Transaksi, pk=pk)

    # generate QR code
    qr = qrcode.make(transaksi.item.kode_barang)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    return render(request, "warehouse/print_keluar.html", {
        "transaksi": transaksi,
        "qr_code": qr_base64,
    })


from datetime import date
from django.utils import timezone
from django.db.models import Sum, Case, When, Value, IntegerField
from django.db.models.functions import ExtractMonth
from django.db.models import Count



from django.http import HttpResponse
from django.template.loader import render_to_string
from .models import Barang, Transaksi

# Export
from xhtml2pdf import pisa
import io
import openpyxl
from openpyxl.utils import get_column_letter

def _filter_transaksi(request, base_qs=None):
    """
    period = date|month|year
    - date  : ?start=YYYY-MM-DD&end=YYYY-MM-DD
    - month : ?year=YYYY&month=MM
    - year  : ?year=YYYY
    """
    if base_qs is None:
        base_qs = Transaksi.objects.select_related('item')

    period = request.GET.get('period', 'date')
    qs = base_qs

    if period == 'date':
        start = request.GET.get('start')
        end = request.GET.get('end')
        if start:
            qs = qs.filter(created_at__date__gte=start)
        if end:
            qs = qs.filter(created_at__date__lte=end)

    elif period == 'month':
        year = request.GET.get('year')
        month = request.GET.get('month')
        if year:
            qs = qs.filter(created_at__year=year)
        if month:
            qs = qs.filter(created_at__month=month)

    elif period == 'year':
        year = request.GET.get('year')
        if year:
            qs = qs.filter(created_at__year=year)

    return qs

from django.core.paginator import Paginator

def report_transaksi(request):
    qs = _filter_transaksi(request)

    total_in = qs.filter(jenis=Transaksi.IN).aggregate(s=Sum('qty'))['s'] or 0
    total_out = qs.filter(jenis=Transaksi.OUT).aggregate(s=Sum('qty'))['s'] or 0

    # Data chart per-bulan untuk tahun yang dipilih (default: tahun sekarang)
    year = int(request.GET.get('year') or timezone.now().year)
    monthly = (
        Transaksi.objects
        .filter(created_at__year=year)
        .annotate(m=ExtractMonth('created_at'))
        .values('m', 'jenis')
        .annotate(total=Sum('qty'))
        .order_by('m')
    )

    in_series = [0]*12
    out_series = [0]*12
    for row in monthly:
        idx = (row['m'] or 1) - 1
        if row['jenis'] == Transaksi.IN:
            in_series[idx] = row['total'] or 0
        else:
            out_series[idx] = row['total'] or 0

    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(qs.order_by('-created_at'), 10)  # tampil 20 per halaman
    page_obj = paginator.get_page(page_number)

    context = {
        'total_in': total_in,
        'total_out': total_out,
        'year': year,
        'in_series': in_series,
        'out_series': out_series,
        'transaksi_list': page_obj,   # diganti page_obj
    }
    return render(request, 'warehouse/report_transaksi.html', context)


from django.core.paginator import Paginator

def report_stok(request):
    qs = _filter_transaksi(request)
    agg = (qs.values('item_id', 'item__kode_barang', 'item__nama_barang')
             .annotate(
                 total_in=Sum(Case(When(jenis=Transaksi.IN, then='qty'),
                                   default=Value(0), output_field=IntegerField())),
                 total_out=Sum(Case(When(jenis=Transaksi.OUT, then='qty'),
                                    default=Value(0), output_field=IntegerField()))
             ))

    # tambah total akhir
    data = []
    for r in agg:
        r['total_qty'] = (r['total_in'] or 0) - (r['total_out'] or 0)
        data.append(r)

    # --- pagination di sini ---
    paginator = Paginator(data, 10)  # tampil 10 item per halaman
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, 'warehouse/report_stok.html', {
        'page_obj': page_obj
    })

# def report_stok(request):
#     qs = _filter_transaksi(request)
#     agg = (qs.values('item_id', 'item__kode_barang', 'item__nama_barang')
#              .annotate(
#                  total_in=Sum(Case(When(jenis=Transaksi.IN, then='qty'),
#                                    default=Value(0), output_field=IntegerField())),
#                  total_out=Sum(Case(When(jenis=Transaksi.OUT, then='qty'),
#                                     default=Value(0), output_field=IntegerField()))
#              ))

#     # tambah total akhir
#     data = []
#     for r in agg:
#         r['total_qty'] = (r['total_in'] or 0) - (r['total_out'] or 0)
#         data.append(r)

#     return render(request, 'warehouse/report_stok.html', {'rows': data})


def export_transaksi_pdf(request):
    qs = _filter_transaksi(request).order_by('-created_at')
    total_in = qs.filter(jenis=Transaksi.IN).aggregate(s=Sum('qty'))['s'] or 0
    total_out = qs.filter(jenis=Transaksi.OUT).aggregate(s=Sum('qty'))['s'] or 0

    html = render_to_string('warehouse/export/transaksi_pdf.html', {
        'rows': qs,
        'total_in': total_in,
        'total_out': total_out,
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="report-transaksi.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    return response


def export_transaksi_xlsx(request):
    qs = _filter_transaksi(request).order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transaksi'

    ws.append(['Tanggal', 'Jenis', 'Kode', 'Nama', 'Qty', 'Status'])
    for t in qs:
        ws.append([
            timezone.localtime(t.created_at).strftime('%Y-%m-%d %H:%M'),
            t.get_jenis_display(),
            t.item.kode_barang,
            t.item.nama_barang,
            t.qty,
            t.get_status_barang_display(),
        ])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="report-transaksi.xlsx"'
    wb.save(response)
    return response

def export_stok_pdf(request):
    qs = _filter_transaksi(request)
    agg = (qs.values('item__kode_barang', 'item__nama_barang')
             .annotate(
                 total_in=Sum(Case(When(jenis=Transaksi.IN, then='qty'),
                                   default=Value(0), output_field=IntegerField())),
                 total_out=Sum(Case(When(jenis=Transaksi.OUT, then='qty'),
                                    default=Value(0), output_field=IntegerField()))
             ))
    rows = []
    for r in agg:
        rows.append({
            'kode': r['item__kode_barang'],
            'nama': r['item__nama_barang'],
            'in': r['total_in'] or 0,
            'out': r['total_out'] or 0,
            'total': (r['total_in'] or 0) - (r['total_out'] or 0),
        })

    html = render_to_string('warehouse/export/stok_pdf.html', {'rows': rows})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="report-stok.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    return response


def export_stok_xlsx(request):
    qs = _filter_transaksi(request)
    agg = (qs.values('item__kode_barang', 'item__nama_barang')
             .annotate(
                 total_in=Sum(Case(When(jenis=Transaksi.IN, then='qty'),
                                   default=Value(0), output_field=IntegerField())),
                 total_out=Sum(Case(When(jenis=Transaksi.OUT, then='qty'),
                                    default=Value(0), output_field=IntegerField()))
             ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stok'
    ws.append(['Kode', 'Nama', 'Total In', 'Total Out', 'Total Qty'])
    for r in agg:
        total_in = r['total_in'] or 0
        total_out = r['total_out'] or 0
        ws.append([
            r['item__kode_barang'],
            r['item__nama_barang'],
            total_in, total_out, total_in - total_out
        ])
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="report-stok.xlsx"'
    wb.save(response)
    return response

from django.core.paginator import Paginator
from PIL import Image
from pyzbar.pyzbar import decode

def upload_qr(request):
    qr_result = None
    if request.method == "POST" and request.FILES.get("qr_file"):
        qr_file = request.FILES["qr_file"]
        img = Image.open(qr_file)
        decoded = decode(img)
        if decoded:
            qr_result = decoded[0].data.decode("utf-8")
        else:
            qr_result = "QR tidak terbaca"

    # ambil ulang data transaksi masuk
    masuk_qs = Transaksi.objects.filter(jenis=Transaksi.IN).order_by("-created_at")
    paginator = Paginator(masuk_qs, 10)  # misal 10 per page
    page_number = request.GET.get("page")
    masuk_list = paginator.get_page(page_number)

    return render(request, "warehouse/barang_masuk.html", {
        "masuk_list": masuk_list,
        "qr_result": qr_result
    })

# apps.py atau signals.py
from django.db.models.signals import post_save
from django.dispatch import receiver
from .models import Barang, Transaksi
from django.contrib.auth import get_user_model
User = get_user_model()
@receiver(post_save, sender=Barang)

def create_initial_transaction(sender, instance, created, **kwargs):
    if created:  # hanya saat barang baru dibuat
        default_user = User.objects.first()
        Transaksi.objects.create(
            item=instance,              # ganti ke 'item'
            qty=instance.qty_total,     # qty awal dari barang
            jenis=Transaksi.IN,
            created_by= default_user
        )

